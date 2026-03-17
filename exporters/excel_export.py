"""
excel_export.py
Exports a CostEstimate to a 5-tab Excel workbook.
Functional-first formatting. Python is the source of truth.
Tabs: Inputs | Assumptions | Outputs | Sensitivity | Notes
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from engine.models import CostEstimate
from engine.assumptions import get_full_assumptions


# ─── STYLE CONSTANTS ──────────────────────────────────────────────────────────

INK       = "1A1A1A"
WHITE     = "FFFFFF"
ACCENT    = "C94A1E"
LIGHT_ROW = "F9F8F5"
RULE      = "D0CFC9"

def _header_font(bold=True):
    return Font(name="Calibri", bold=bold, color=WHITE, size=10)

def _body_font(bold=False, color=INK):
    return Font(name="Calibri", bold=bold, color=color, size=10)

def _header_fill():
    return PatternFill("solid", fgColor=INK)

def _accent_fill():
    return PatternFill("solid", fgColor=ACCENT)

def _light_fill():
    return PatternFill("solid", fgColor=LIGHT_ROW)

def _thin_border():
    s = Side(style="thin", color=RULE)
    return Border(bottom=s)

def _write_header_row(ws, row: int, values: list[str], col_start: int = 1):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=val)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="left", vertical="center")

def _write_section_label(ws, row: int, label: str, col: int = 1):
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = Font(name="Calibri", bold=True, color=ACCENT, size=9)

def _write_kv(ws, row: int, key: str, value, bold_value=False, col: int = 1):
    k = ws.cell(row=row, column=col, value=key)
    k.font = _body_font(bold=True)
    v = ws.cell(row=row, column=col + 1, value=value)
    v.font = _body_font(bold=bold_value)
    v.border = _thin_border()

def _auto_width(ws, min_width=12, max_width=60):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ─── TAB BUILDERS ─────────────────────────────────────────────────────────────

def _build_inputs_tab(ws, estimate: CostEstimate, inputs_raw: dict):
    ws.title = "Inputs"
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="SHOULD-COST MODEL — INPUTS").font = Font(
        name="Calibri", bold=True, size=13, color=INK)
    r += 1
    ws.cell(row=r, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _body_font(color="888888")
    r += 2

    _write_section_label(ws, r, "PART IDENTIFICATION")
    r += 1
    _write_kv(ws, r, "Part ID",          estimate.part_id);          r += 1
    _write_kv(ws, r, "Description",      estimate.part_description); r += 1
    _write_kv(ws, r, "Program",          inputs_raw.get("program", "—")); r += 1
    r += 1

    _write_section_label(ws, r, "MATERIAL & GEOMETRY")
    r += 1
    _write_kv(ws, r, "Material",         estimate.material);          r += 1
    _write_kv(ws, r, "Finished Weight (lb)", inputs_raw.get("finished_weight_lb")); r += 1
    _write_kv(ws, r, "Complexity Tier",  inputs_raw.get("complexity_tier")); r += 1
    _write_kv(ws, r, "Tolerance Tier",   inputs_raw.get("tolerance_tier")); r += 1
    r += 1

    _write_section_label(ws, r, "VOLUME & REGION")
    r += 1
    _write_kv(ws, r, "Annual Volume",    estimate.annual_volume);  r += 1
    _write_kv(ws, r, "Batch Size",       estimate.batch_size);     r += 1
    _write_kv(ws, r, "Region",           estimate.region);         r += 1
    r += 1

    _write_section_label(ws, r, "OVERRIDES")
    r += 1
    _write_kv(ws, r, "Machining Hours Override",
              inputs_raw.get("machining_hours_override", "None — model estimate used")); r += 1
    _write_kv(ws, r, "Setup Hours Override",
              inputs_raw.get("setup_hours_override", "None — model estimate used")); r += 1
    r += 1

    _write_section_label(ws, r, "OUTSIDE PROCESSES")
    r += 1
    procs = inputs_raw.get("outside_processes", [])
    if procs:
        for p in procs:
            _write_kv(ws, r, "", p); r += 1
    else:
        _write_kv(ws, r, "", "None"); r += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30


def _build_assumptions_tab(ws, estimate: CostEstimate):
    ws.title = "Assumptions"
    ws.sheet_view.showGridLines = False

    au = estimate.assumptions_used
    r = 1
    ws.cell(row=r, column=1, value="ASSUMPTIONS USED IN THIS ESTIMATE").font = Font(
        name="Calibri", bold=True, size=13, color=INK)
    r += 1
    ws.cell(row=r, column=1,
            value="All values reflect mid (baseline) assumptions unless noted. "
                  "Modify assumptions_default.json to change defaults.").font = _body_font(color="888888")
    r += 2

    _write_section_label(ws, r, "MATERIAL")
    r += 1
    _write_kv(ws, r, "Material Price ($/lb)",  f"${au.material_price_per_lb:.2f}"); r += 1
    _write_kv(ws, r, "Buy-to-Fly Ratio",       f"{au.buy_to_fly_ratio:.2f}x");      r += 1
    r += 1

    _write_section_label(ws, r, "MACHINING")
    r += 1
    _write_kv(ws, r, "Machining Hours",        f"{au.machining_hours:.2f} hrs");    r += 1
    _write_kv(ws, r, "Setup Hours",            f"{au.setup_hours:.2f} hrs");         r += 1
    _write_kv(ws, r, "Machine Rate ($/hr)",    f"${au.machine_rate_per_hr:.2f}");    r += 1
    r += 1

    _write_section_label(ws, r, "COST FACTORS")
    r += 1
    _write_kv(ws, r, "Overhead Rate",          f"{au.overhead_rate * 100:.0f}%");    r += 1
    _write_kv(ws, r, "Supplier Margin Rate",   f"{au.margin_rate * 100:.0f}%");      r += 1
    _write_kv(ws, r, "Scrap Factor",           f"{au.scrap_factor * 100:.1f}%");     r += 1
    r += 1

    if au.outside_process_costs:
        _write_section_label(ws, r, "OUTSIDE PROCESS COSTS (PER PIECE, MID)")
        r += 1
        for proc, cost in au.outside_process_costs.items():
            _write_kv(ws, r, proc.replace("_", " ").title(), f"${cost:.2f}"); r += 1
    else:
        _write_kv(ws, r, "Outside Processes", "None"); r += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22


def _build_outputs_tab(ws, estimate: CostEstimate):
    ws.title = "Outputs"
    ws.sheet_view.showGridLines = False

    bd = estimate.breakdown
    pb = estimate.price_band
    r = 1

    ws.cell(row=r, column=1, value="SHOULD-COST RESULTS").font = Font(
        name="Calibri", bold=True, size=13, color=INK)
    r += 2

    # Price band — prominent
    _write_section_label(ws, r, "PRICE BAND")
    r += 1
    _write_header_row(ws, r, ["Scenario", "Unit Price", "Notes"])
    r += 1

    band_rows = [
        ("Low (Optimistic)",   pb.low,  "Best-case material, rates, margin"),
        ("Mid (Baseline)",     pb.mid,  "Standard assumptions — primary reference"),
        ("High (Conservative)",pb.high, "Worst-case material, rates, margin"),
    ]
    for i, (label, price, note) in enumerate(band_rows):
        fill = _light_fill() if i % 2 == 0 else None
        for col, val in enumerate([label, f"${price:,.2f}", note], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _body_font(bold=(label == "Mid (Baseline)"))
            if fill:
                cell.fill = fill
        r += 1
    r += 1

    # Cost breakdown
    _write_section_label(ws, r, "COST BREAKDOWN — MID SCENARIO (PER PIECE)")
    r += 1
    _write_header_row(ws, r, ["Cost Element", "Amount", "% of Unit Price"])
    r += 1

    breakdown_rows = [
        ("Material",               bd.material_cost),
        ("Machining",              bd.machining_cost),
        ("Setup (amortized)",      bd.setup_cost_per_piece),
        ("Outside Processes",      bd.outside_process_cost),
        ("Scrap",                  bd.scrap_cost),
        ("Overhead",               bd.overhead_cost),
        ("Supplier Margin",        bd.supplier_margin),
    ]
    total_shown = sum(v for _, v in breakdown_rows)
    for i, (label, amount) in enumerate(breakdown_rows):
        pct = f"{(amount / bd.unit_price_mid * 100):.1f}%" if bd.unit_price_mid > 0 else "—"
        fill = _light_fill() if i % 2 == 0 else None
        for col, val in enumerate([label, f"${amount:,.2f}", pct], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _body_font()
            if fill:
                cell.fill = fill
        r += 1

    # Total row
    for col, val in enumerate(["UNIT PRICE (MID)", f"${bd.unit_price_mid:,.2f}", "100%"], start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill = _header_fill()
    r += 2

    # Confidence
    _write_section_label(ws, r, "CONFIDENCE")
    r += 1
    _write_kv(ws, r, "Confidence Level", estimate.confidence.value, bold_value=True); r += 1
    for note in estimate.confidence_notes:
        _write_kv(ws, r, "", note); r += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 36


def _build_sensitivity_tab(ws, estimate: CostEstimate):
    ws.title = "Sensitivity"
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="SENSITIVITY ANALYSIS").font = Font(
        name="Calibri", bold=True, size=13, color=INK)
    r += 1
    ws.cell(row=r, column=1,
            value="Impact on mid unit price if each variable increases by +10%. Top 3 drivers shown.").font = _body_font(color="888888")
    r += 2

    _write_header_row(ws, r, ["Variable", "Price Delta ($)", "Price Delta (%)", "Interpretation"])
    r += 1

    interp = {
        "material_price":  "Material is a primary cost lever — sourcing raw material competitively matters.",
        "machining_hours": "Cycle time is a significant driver — complexity and setup efficiency matter.",
        "machine_rate":    "Machine rate variation impacts cost — regional rate differences are meaningful.",
        "scrap_factor":    "Scrap rate has measurable impact — process capability and yield are cost factors.",
    }

    for i, driver in enumerate(estimate.sensitivity):
        fill = _light_fill() if i % 2 == 0 else None
        note = interp.get(driver.variable, "")
        for col, val in enumerate(
            [driver.variable.replace("_", " ").title(),
             f"+${driver.delta_dollar:,.2f}",
             f"+{driver.delta_pct:.1f}%",
             note],
            start=1
        ):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _body_font()
            if fill:
                cell.fill = fill
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="Note: Sensitivity shows directional magnitude only. "
                  "All variables can move simultaneously in adverse conditions.").font = _body_font(color="888888")

    _auto_width(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 55


def _build_notes_tab(ws, estimate: CostEstimate):
    ws.title = "Notes"
    ws.sheet_view.showGridLines = False

    r = 1
    ws.cell(row=r, column=1, value="NOTES & NARRATIVE").font = Font(
        name="Calibri", bold=True, size=13, color=INK)
    r += 2

    if estimate.ai_narrative:
        _write_section_label(ws, r, "AI-GENERATED SOURCING NARRATIVE")
        r += 1
        ws.cell(row=r, column=1,
                value="Scoring is fully deterministic. AI rewrites computed drivers into plain language only.").font = _body_font(color="888888")
        r += 1
        narrative_cell = ws.cell(row=r, column=1, value=estimate.ai_narrative)
        narrative_cell.font = _body_font()
        narrative_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 120
        r += 2
    else:
        ws.cell(row=r, column=1,
                value="No AI narrative generated. Set ANTHROPIC_API_KEY to enable.").font = _body_font(color="888888")
        r += 2

    _write_section_label(ws, r, "ANALYST NOTES")
    r += 1
    note_cell = ws.cell(row=r, column=1, value=estimate.notes if hasattr(estimate, "notes") else "")
    note_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 60
    r += 2

    _write_section_label(ws, r, "MODEL LIMITATIONS")
    r += 1
    limitations = [
        "This model estimates machined parts only. Castings, forgings, and composites are out of scope.",
        "Buy-to-fly ratios are table-driven by material + complexity tier. Geometry-specific ratios require V1.5 STEP ingestion.",
        "Machining hours are estimated from complexity and tolerance tiers, not feature counts or cycle time analysis.",
        "Outside process costs are industry benchmarks. Actual supplier quotes should replace these for high-value parts.",
        "Price band does not account for tooling amortization, NRE, or first-article costs.",
        "Python is the source of truth. Values modified in this Excel file do not feed back into the model.",
    ]
    for lim in limitations:
        ws.cell(row=r, column=1, value=f"— {lim}").font = _body_font(color="888888")
        r += 1

    ws.column_dimensions["A"].width = 90


# ─── MAIN EXPORT ──────────────────────────────────────────────────────────────

def export_to_excel(
    estimate: CostEstimate,
    inputs_raw: dict,
    output_path: str | Path | None = None,
) -> Path:
    """
    Export a CostEstimate to a 5-tab Excel workbook.

    Args:
        estimate:    CostEstimate from estimator.py (optionally with ai_narrative from explain.py)
        inputs_raw:  Raw dict of PartInputs fields for the Inputs tab
        output_path: Destination path. Defaults to ./{part_id}_should_cost.xlsx

    Returns:
        Path to the created file.
    """
    if output_path is None:
        safe_id = estimate.part_id.replace("/", "-").replace(" ", "_")
        output_path = Path(f"{safe_id}_should_cost.xlsx")
    output_path = Path(output_path)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_inputs      = wb.create_sheet("Inputs")
    ws_assumptions = wb.create_sheet("Assumptions")
    ws_outputs     = wb.create_sheet("Outputs")
    ws_sensitivity = wb.create_sheet("Sensitivity")
    ws_notes       = wb.create_sheet("Notes")

    _build_inputs_tab(ws_inputs, estimate, inputs_raw)
    _build_assumptions_tab(ws_assumptions, estimate)
    _build_outputs_tab(ws_outputs, estimate)
    _build_sensitivity_tab(ws_sensitivity, estimate)
    _build_notes_tab(ws_notes, estimate)

    wb.save(output_path)
    return output_path
